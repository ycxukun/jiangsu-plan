#!/usr/bin/env python3
"""Build the source-audited 2026 undergraduate-major catalog for the preview app."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


OFFICIAL_NOTICE_URL = (
    "https://hudong.moe.gov.cn/srcsite/A08/moe_1034/s3882/202604/"
    "t20260427_1434931.html"
)
OFFICIAL_PDF_URL = (
    "https://www.moe.gov.cn/srcsite/A08/moe_1034/s3882/202604/"
    "W020260427440749576927.pdf"
)
OFFICIAL_SOURCE_NAME = "普通高等学校本科专业目录（2026年）"
SUMMARY_SOURCE_NAME = "专业详解.xlsx"
CLASS_SOURCE_NAME = "专业大类详解"
COURSE_SOURCE_NAME = "大学专业课程.xlsx"
CAREER_SOURCE_NAME = "专业就业去向.xlsx"
UNIVERSITY_SOURCE_NAME = "专业院校宝典.xlsx"

ACCENTS = {
    "01": "blue",
    "02": "amber",
    "03": "red",
    "04": "green",
    "05": "purple",
    "06": "orange",
    "07": "cyan",
    "08": "blue",
    "09": "green",
    "10": "cyan",
    "12": "amber",
    "13": "purple",
    "14": "red",
}

TRAIT_WORDS = {
    "数学": ("数学", "建模", "统计", "计算"),
    "编程": ("编程", "程序", "软件", "算法", "计算机", "代码"),
    "物理": ("物理", "力学", "电气", "电子", "机械", "能源"),
    "实验": ("实验", "化学", "生物", "医学", "材料"),
    "深造": ("科研", "研究", "深造", "学术", "理论"),
    "项目": ("工程", "项目", "设计", "制造", "实践"),
    "阅读": ("阅读", "写作", "语言", "法学", "文学", "历史", "哲学"),
    "数字": ("数据", "数字", "信息", "智能", "统计", "金融"),
}

NUMBERED_PREFIX = re.compile(r"^[0-9一二三四五六七八九十百〇零]+[.、．]\s*")
H1_CLASS_RE = re.compile(r"^#\s+\*{0,2}([0-9]{4})\s+([^｜|\n*]+)", re.M)
H2_RE = re.compile(r"^##\s+(.+?)\s*$", re.M)
H2_MAJOR_RE = re.compile(
    r"^##\s+\*{0,2}(?:[一二三四五六七八九十百〇零]+、)?\s*"
    r"([0-9]{6}[A-Z]{0,2})\s+([^\n*]+)",
    re.M,
)

# These mappings are explicit professional-class mappings. Do not replace them
# with fuzzy title matching: a similarly named major may belong to another class.
CATEGORY_COURSE_SHEET = {
    "0201": "经济金融类",
    "0203": "经济金融类",
    "0301": "法学类",
    "0701": "数学类",
    "0702": "物理类",
    "0712": "统计学类",
    "0802": "机械类",
    "0803": "仪器类",
    "0804": "材料类",
    "0805": "能动类",
    "0806": "电气类",
    "0807": "电子类",
    "0808": "控制类",
    "0809": "计算机",
    "0819": "船舶类",
    "0820": "航空航天类",
    "1002": "临床医学类",
}

MAJOR_COURSE_SHEET = {
    "通信工程": "通信类",
    "电信工程及管理": "通信类",
    "会计学": "财会类",
    "财务管理": "财会类",
    "审计学": "财会类",
    "财务会计教育": "财会类",
}

CATEGORY_CAREER_SHEET = {
    "0201": "经",
    "0203": "金",
    "0301": "法",
    "0502": "外",
    "0701": "数统",
    "0703": "化",
    "0706": "气",
    "0709": "地",
    "0710": "生",
    "0712": "数统",
    "0801": "力",
    "0802": "械",
    "0803": "仪",
    "0804": "材",
    "0805": "能动",
    "0806": "电气",
    "0807": "电子",
    "0808": "自",
    "0809": "计",
    "0810": "土",
    "0811": "水",
    "0813": "化工",
    "0814": "地",
    "0818": "交",
    "0819": "船",
    "0820": "航",
    "0821": "兵",
    "0825": "环",
    "0826": "医工",
    "0827": "食",
    "0829": "安",
    "0904": "兽",
    "1002": "临床",
    "1005": "中",
    "1007": "药",
    "1010": "医技",
}

MAJOR_CAREER_SHEET = {
    "通信工程": "通",
    "电信工程及管理": "通",
    "光电信息科学与工程": "光",
    "新能源科学与工程": "新",
    "会计学": "财",
    "财务管理": "财",
    "审计学": "财",
    "工商管理": "管",
    "市场营销": "管",
    "人力资源管理": "管",
}

CATEGORY_UNIVERSITY_SHEET = {
    "0701": "理学工学:数学",
    "0702": "理学工学:物理",
    "0703": "理学工学:化学",
    "0705": "理学工学:地理",
    "0707": "理学工学:海洋科学",
    "0708": "理学工学:地球物理学",
    "0709": "理学工学:地质学",
    "0710": "理学工学:生物学",
    "0711": "理学工学:系统科学",
    "0712": "统计学",
    "0802": "机械",
    "0803": "仪器",
    "0804": "材料",
    "0805": "能源动力",
    "0807": "电子信息类",
    "0808": "自动化",
    "0809": "计算机",
    "0818": "交通运输",
    "0820": "航空航天类",
    "0826": "生医工",
}


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u3000", " ").strip()
    text = text.replace("**", "").replace("`", "")
    return re.sub(r"\s+", " ", text)


def shorten(value: object, limit: int = 260) -> str:
    text = clean_text(value)
    if len(text) <= limit:
        return text
    return text[: limit - 1].rstrip("，,；;。 ") + "…"


def unique(values: Iterable[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = clean_text(value)
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result


def split_keywords(value: object) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    values = [part.strip(" ，,、；;。") for part in re.split(r"[、，,；;]+", text)]
    return [part for part in values if part][:6]


def split_risks(value: object, major_name: str) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    text = re.sub(rf"^{re.escape(major_name)}\s*[,，:：]\s*", "", text)
    parts = [part.strip(" ，,、；;。") for part in re.split(r"[。；;]+", text)]
    return [shorten(part, 180) for part in parts if len(part) >= 8][:4]


def derive_traits(*values: object) -> list[str]:
    haystack = " ".join(clean_text(value) for value in values)
    return [
        trait
        for trait, words in TRAIT_WORDS.items()
        if any(word in haystack for word in words)
    ]


def degree_from_note(note: str) -> str:
    if not note:
        return "以院校培养方案为准"
    match = re.search(r"(?:授予|可授)(.+?)学士学位", note)
    if match:
        return match.group(1).strip("，,；;。 ")
    return "以院校培养方案为准"


def sentence_candidates(text: str) -> list[str]:
    normalized = clean_text(text)
    parts = re.split(r"(?<=[。！？])\s*|[；;]\s*", normalized)
    return [part.strip(" ，,；;。") for part in parts if 12 <= len(part.strip()) <= 280]


def numbered_points(text: str, limit: int = 5) -> list[str]:
    points: list[str] = []
    for paragraph in re.split(r"\n\s*\n", text):
        cleaned = clean_text(paragraph)
        if not cleaned or cleaned.startswith("##"):
            continue
        for sentence in sentence_candidates(cleaned):
            sentence = re.sub(
                r"^(?:第一|第二|第三|第四|第五|第六|第七|一是|二是|三是|四是)[，,:：]\s*",
                "",
                sentence,
            )
            if sentence and sentence not in points:
                points.append(shorten(sentence, 180))
            if len(points) >= limit:
                return points
    return points


def section_by_heading(body: str, predicate) -> str:
    headings = list(H2_RE.finditer(body))
    for index, heading in enumerate(headings):
        title = clean_text(heading.group(1))
        if not predicate(title):
            continue
        end = headings[index + 1].start() if index + 1 < len(headings) else len(body)
        return body[heading.end() : end].strip()
    return ""


def informative_paragraphs(text: str, limit: int = 5) -> list[str]:
    result: list[str] = []
    for paragraph in re.split(r"\n\s*\n", text):
        cleaned = clean_text(paragraph)
        if not cleaned or cleaned.startswith("#") or len(cleaned) < 18:
            continue
        if cleaned.startswith(("下面", "这是", "在《", "这套分类")):
            continue
        result.append(shorten(cleaned, 280))
        if len(result) >= limit:
            break
    return result


def risk_points(text: str, limit: int = 4) -> list[str]:
    terms = ("风险", "问题", "劣势", "谨慎", "慎选", "不适合", "不直接", "岗位少", "门槛", "分化")
    result: list[str] = []
    for sentence in sentence_candidates(text):
        if any(term in sentence for term in terms):
            result.append(shorten(sentence, 180))
        if len(result) >= limit:
            break
    return unique(result)


def parse_class_details(path: Path | None) -> tuple[dict[str, dict[str, object]], dict[str, dict[str, object]], dict[str, object]]:
    if not path:
        return {}, {}, {"duplicateClassCodes": [], "extraMajorCodes": []}
    text = path.read_text(encoding="utf-8")
    class_matches = list(H1_CLASS_RE.finditer(text))
    raw_classes: dict[str, list[tuple[str, str]]] = {}
    major_details: dict[str, dict[str, object]] = {}

    for index, match in enumerate(class_matches):
        code = match.group(1)
        name = clean_text(match.group(2))
        end = class_matches[index + 1].start() if index + 1 < len(class_matches) else len(text)
        body = text[match.end() : end]
        raw_classes.setdefault(code, []).append((name, body))

    class_details: dict[str, dict[str, object]] = {}
    for code, versions in raw_classes.items():
        name, body = max(versions, key=lambda item: len(item[1]))
        overview_body = section_by_heading(body, lambda title: "总论" in title)
        overview = informative_paragraphs(overview_body or body, 3)
        suitable_body = section_by_heading(
            body, lambda title: "适合什么样" in title and "不适合" not in title
        )
        unsuitable_body = section_by_heading(body, lambda title: "不适合什么样" in title)
        career_body = section_by_heading(
            body, lambda title: "就业真相" in title or "就业出口" in title
        )
        talk_body = section_by_heading(body, lambda title: "给家长的简化表达" in title)
        conclusion_body = section_by_heading(body, lambda title: "最终结论" in title)

        class_details[code] = {
            "code": code,
            "name": name,
            "overview": overview,
            "suitable": numbered_points(suitable_body),
            "unsuitable": numbered_points(unsuitable_body),
            "graduate": informative_paragraphs(career_body, 3),
            "risks": risk_points(career_body + "\n" + conclusion_body),
            "talkTrack": informative_paragraphs(talk_body, 3),
        }

        major_matches = list(H2_MAJOR_RE.finditer(body))
        all_h2_matches = list(H2_RE.finditer(body))
        for major_match in major_matches:
            major_code = major_match.group(1)
            major_name = clean_text(major_match.group(2))
            major_end = next(
                (heading.start() for heading in all_h2_matches if heading.start() > major_match.start()),
                len(body),
            )
            major_body = body[major_match.end() : major_end]
            paragraphs = informative_paragraphs(major_body, 8)
            detail = {
                "code": major_code,
                "name": major_name,
                "paragraphs": paragraphs,
                "academic": paragraphs[0] if paragraphs else "",
                "parent": paragraphs[1] if len(paragraphs) > 1 else (paragraphs[0] if paragraphs else ""),
                "suitable": [
                    shorten(sentence, 180)
                    for sentence in sentence_candidates(major_body)
                    if "适合" in sentence and "不适合" not in sentence
                ][:4],
                "unsuitable": [
                    shorten(sentence, 180)
                    for sentence in sentence_candidates(major_body)
                    if any(term in sentence for term in ("不适合", "慎选", "谨慎"))
                ][:4],
                "graduate": [
                    shorten(sentence, 180)
                    for sentence in sentence_candidates(major_body)
                    if any(term in sentence for term in ("就业", "深造", "读研", "考公", "考编"))
                ][:3],
                "risks": risk_points(major_body),
            }
            previous = major_details.get(major_code)
            if previous is None or len(clean_text(detail["academic"])) > len(clean_text(previous["academic"])):
                major_details[major_code] = detail

    audit = {
        "duplicateClassCodes": sorted(code for code, values in raw_classes.items() if len(values) > 1),
        "classHeadingCount": sum(len(values) for values in raw_classes.values()),
        "uniqueClassCount": len(class_details),
        "specificMajorHeadingCount": len(major_details),
    }
    return class_details, major_details, audit


def load_local_summaries(workbook_path: Path) -> tuple[dict[str, dict[str, object]], dict[str, object]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["总表"]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean_text(value) for value in next(rows)]
    result: dict[str, dict[str, object]] = {}
    duplicate_names: list[str] = []
    for values in rows:
        row = dict(zip(headers, values))
        name = clean_text(row.get("专业名称"))
        if not name:
            continue
        if name in result:
            duplicate_names.append(name)
            continue
        result[name] = row
    workbook.close()
    return result, {
        "sourceRowCount": len(result),
        "duplicateNames": sorted(set(duplicate_names)),
    }


def load_course_templates(path: Path | None) -> dict[str, list[dict[str, object]]]:
    if not path:
        return {}
    workbook = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, list[dict[str, object]]] = {}
    for sheet in workbook.worksheets:
        if sheet.title in {"带电工科", "力学工科"}:
            continue
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [clean_text(value) for value in rows[0]]
        modules: list[dict[str, object]] = []
        for values in rows[1:]:
            row = dict(zip(headers, values))
            name = clean_text(row.get("模块名称"))
            if not name:
                continue
            name = NUMBERED_PREFIX.sub("", name)
            courses_text = clean_text(
                row.get("核心课程清单 (部分展示)")
                or row.get("核心理论/工具")
                or row.get("核心研究对象")
            )
            value = clean_text(
                row.get("学科价值与定位")
                or row.get("典型工程应用")
                or row.get("核心研究对象")
            )
            courses = [
                part.strip()
                for part in re.split(r"[、，,；;]", courses_text)
                if part.strip()
            ]
            modules.append({"name": name, "courses": courses[:16], "value": value})
        if modules:
            result[sheet.title] = modules
    workbook.close()
    return result


def load_career_templates(path: Path | None) -> dict[str, list[dict[str, str]]]:
    if not path:
        return {}
    workbook = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, list[dict[str, str]]] = {}
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [clean_text(value) for value in rows[0]]
        careers: list[dict[str, str]] = []
        for index, values in enumerate(rows[1:], start=1):
            row = dict(zip(headers, values))
            serial = clean_text(row.get("序号"))
            if not serial or not re.fullmatch(r"\d+", serial):
                continue
            detail_values = rows[index + 2] if index + 2 < len(rows) else ()
            detail = dict(zip(headers, detail_values))
            work = clean_text(row.get("工作简要与核心职责"))
            detail_work = clean_text(detail.get("工作简要与核心职责"))
            if detail_work:
                work = f"{work}：{detail_work}" if work else detail_work
            skills_key = next((header for header in headers if "核心技能" in header), "")
            skills = clean_text(row.get(skills_key))
            detail_skills = clean_text(detail.get(skills_key))
            if detail_skills:
                skills = f"{skills}；{detail_skills}" if skills else detail_skills
            careers.append(
                {
                    "name": clean_text(row.get("细分方向")),
                    "attribute": clean_text(row.get("赛道属性")),
                    "work": work,
                    "employers": clean_text(row.get("代表单位")),
                    "skills": skills,
                    "upside": clean_text(row.get("核心优势 (Pros)")),
                    "risk": clean_text(row.get("职业痛点 (Cons)")),
                }
            )
        if careers:
            result[sheet.title] = careers
    workbook.close()
    return result


def school_level(label: str) -> str:
    if "985" in label:
        return "top"
    if any(term in label for term in ("211", "双一流", "国重点", "特色")):
        return "feature"
    return "local"


def load_school_table(sheet) -> list[dict[str, str]]:
    rows = list(sheet.iter_rows(values_only=True))
    header_index = -1
    headers: list[str] = []
    for index, values in enumerate(rows[:8]):
        candidate = [clean_text(value) for value in values]
        if any(value in {"院校名称", "院校"} for value in candidate):
            header_index = index
            headers = candidate
            break
    if header_index < 0:
        return []
    name_key = next((value for value in headers if value in {"院校名称", "院校"}), "")
    province_key = next((value for value in headers if value in {"省份", "所在省"}), "")
    city_key = next((value for value in headers if value == "城市"), "")
    label_key = next((value for value in headers if "标签" in value or "层次" in value), "")
    result: list[dict[str, str]] = []
    for values in rows[header_index + 1 :]:
        row = dict(zip(headers, values))
        name = clean_text(row.get(name_key))
        if not name or name in {item["name"] for item in result}:
            continue
        label = clean_text(row.get(label_key))
        evidence: list[str] = []
        for header in headers:
            if header in {"序号", name_key, province_key, city_key, label_key, "计数"}:
                continue
            value = clean_text(row.get(header))
            if not value:
                continue
            short_header = clean_text(header.replace("\n", ""))
            evidence.append(f"{short_header}{value}" if value != "√" else short_header)
            if len(evidence) >= 3:
                break
        result.append(
            {
                "name": name,
                "province": clean_text(row.get(province_key)) or "—",
                "city": clean_text(row.get(city_key)) or "—",
                "level": school_level(label),
                "strength": "；".join(evidence) or f"{sheet.title}资料收录",
                "note": label or f"来源：{sheet.title}",
            }
        )
    return result[:18]


def load_matrix_school_table(sheet, topic: str) -> list[dict[str, str]]:
    rows = list(sheet.iter_rows(values_only=True))
    for values in rows:
        first = clean_text(values[0] if values else "")
        if first != topic:
            continue
        result: list[dict[str, str]] = []
        cells = [clean_text(value) for value in values]
        for index in range(1, len(cells) - 1, 2):
            grade = cells[index]
            schools = cells[index + 1]
            if not schools:
                continue
            for school in re.split(r"[\n、，,；;]", schools):
                name = clean_text(school)
                if not name or name in {item["name"] for item in result}:
                    continue
                result.append(
                    {
                        "name": name,
                        "province": "—",
                        "city": "—",
                        "level": "feature",
                        "strength": f"{topic}学科评估参考：{grade or '已收录'}",
                        "note": f"来源：{sheet.title}，具体招生专业以当年计划为准",
                    }
                )
        return result[:18]
    return []


def load_university_templates(path: Path | None) -> dict[str, list[dict[str, str]]]:
    if not path:
        return {}
    workbook = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, list[dict[str, str]]] = {}
    for sheet in workbook.worksheets:
        schools = load_school_table(sheet)
        if schools:
            result[sheet.title] = schools
    if "理学工学" in workbook.sheetnames:
        for topic in ("数学", "物理", "化学", "地理", "海洋科学", "地球物理学", "地质学", "生物学", "系统科学"):
            schools = load_matrix_school_table(workbook["理学工学"], topic)
            if schools:
                result[f"理学工学:{topic}"] = schools
    workbook.close()
    return result


def course_sheet_for(record: dict[str, object]) -> str:
    name = clean_text(record.get("name"))
    if name in MAJOR_COURSE_SHEET:
        return MAJOR_COURSE_SHEET[name]
    if clean_text(record.get("categoryCode")) == "0807" and "通信" in name:
        return "通信类"
    return CATEGORY_COURSE_SHEET.get(clean_text(record.get("categoryCode")), "")


def career_sheet_for(record: dict[str, object]) -> str:
    name = clean_text(record.get("name"))
    if name in MAJOR_CAREER_SHEET:
        return MAJOR_CAREER_SHEET[name]
    if clean_text(record.get("categoryCode")) == "0807":
        if "通信" in name:
            return "通"
        if "光电" in name:
            return "光"
    return CATEGORY_CAREER_SHEET.get(clean_text(record.get("categoryCode")), "")


def university_source_for(record: dict[str, object]) -> str:
    return CATEGORY_UNIVERSITY_SHEET.get(clean_text(record.get("categoryCode")), "")


def load_official_records(parsed_path: Path | None, catalog_js_path: Path | None) -> list[dict[str, object]]:
    if parsed_path:
        parsed = json.loads(parsed_path.read_text(encoding="utf-8"))
        return list(parsed["records"])
    if not catalog_js_path:
        raise SystemExit("Provide --parsed or --catalog-js")
    javascript = catalog_js_path.read_text(encoding="utf-8")
    marker = "window.MAJOR_CATALOG_2026 = "
    start = javascript.index(marker) + len(marker)
    end = javascript.index(";\n}());", start)
    return list(json.loads(javascript[start:end]))


def official_identity(record: dict[str, object]) -> dict[str, object]:
    return {
        "name": clean_text(record.get("name")),
        "code": clean_text(record.get("code")),
        "disciplineCode": clean_text(record.get("disciplineCode")),
        "discipline": clean_text(record.get("discipline")),
        "categoryCode": clean_text(record.get("categoryCode")),
        "category": clean_text(record.get("category")) or "目录直列专业",
        "note": clean_text(record.get("note") or record.get("officialNote")),
    }


def normalized_record(
    record: dict[str, object],
    local: dict[str, object] | None,
    class_detail: dict[str, object] | None,
    major_detail: dict[str, object] | None,
    course_templates: dict[str, list[dict[str, object]]],
    career_templates: dict[str, list[dict[str, str]]],
    university_templates: dict[str, list[dict[str, str]]],
) -> tuple[dict[str, object], dict[str, object]]:
    identity = official_identity(record)
    name = str(identity["name"])
    code = str(identity["code"])
    discipline = str(identity["discipline"])
    category = str(identity["category"])
    note = str(identity["note"])

    has_major_detail = local is not None or major_detail is not None
    has_class_detail = class_detail is not None
    content_status = "major" if has_major_detail else ("class" if has_class_detail else "catalog")

    keywords = split_keywords(local.get("核心关键词")) if local else []
    academic = clean_text(local.get("一句话详细解读 (学术/技术定义)")) if local else ""
    parent = clean_text(local.get("通俗/直观理解")) if local else ""
    risks = split_risks(local.get("避坑行业内情信息差"), name) if local else []
    scale_value = local.get("招生计划") if local else None
    try:
        scale = int(float(scale_value)) if scale_value not in (None, "") else None
    except (TypeError, ValueError):
        scale = None

    if major_detail:
        if not academic:
            academic = shorten(major_detail.get("academic"), 280)
        if not parent:
            parent = shorten(major_detail.get("parent"), 260)
        risks = unique([*risks, *major_detail.get("risks", [])])[:4]
    if not academic:
        academic = f"{name}是教育部2026年本科专业目录中{discipline}门类、{category}下的专业。"
    if not parent and class_detail:
        overview = list(class_detail.get("overview", []))
        if overview:
            parent = f"专业类参考：{shorten(overview[0], 230)}"
    if not parent:
        parent = "官方目录已收录该专业；具体课程、培养方向和就业讲解仍待可靠资料补充。"

    suitable = list(major_detail.get("suitable", [])) if major_detail else []
    unsuitable = list(major_detail.get("unsuitable", [])) if major_detail else []
    graduate = list(major_detail.get("graduate", [])) if major_detail else []
    talk_track: list[str] = []
    if class_detail:
        if not suitable:
            suitable = [f"专业类参考：{item}" for item in class_detail.get("suitable", [])]
        if not unsuitable:
            unsuitable = [f"专业类参考：{item}" for item in class_detail.get("unsuitable", [])]
        if not graduate:
            graduate = [f"专业类参考：{item}" for item in class_detail.get("graduate", [])]
        if not risks:
            risks = [f"专业类参考：{item}" for item in class_detail.get("risks", [])]
        talk_track = [f"专业类参考：{item}" for item in class_detail.get("talkTrack", [])]
    if not risks:
        risks = ["专业目录身份已核对，院校培养方向、选科要求和当年招生章程仍需逐校确认。"]

    course_sheet = course_sheet_for(identity)
    career_sheet = career_sheet_for(identity)
    university_source = university_source_for(identity)
    modules = course_templates.get(course_sheet, [])
    careers = career_templates.get(career_sheet, [])
    universities = university_templates.get(university_source, [])

    flag_labels: list[str] = []
    if "T" in code:
        flag_labels.append("特设专业")
    if "K" in code:
        flag_labels.append("国家控制布点专业")

    sources = [OFFICIAL_SOURCE_NAME]
    if local:
        sources.append(SUMMARY_SOURCE_NAME)
    if major_detail or class_detail:
        sources.append(CLASS_SOURCE_NAME)
    if modules:
        sources.append(COURSE_SOURCE_NAME)
    if careers:
        sources.append(CAREER_SOURCE_NAME)
    if universities:
        sources.append(UNIVERSITY_SOURCE_NAME)

    normalized = {
        "id": f"major-{code.lower()}",
        "code": code,
        "name": name,
        "disciplineCode": identity["disciplineCode"],
        "discipline": discipline,
        "categoryCode": identity["categoryCode"],
        "category": category,
        "officialNote": note,
        "officialYear": 2026,
        "flagLabels": flag_labels,
        "contentStatus": content_status,
        "contentLevel": "专业级" if has_major_detail else ("专业类参考" if has_class_detail else "仅目录"),
        "duration": "以院校培养方案为准",
        "degree": degree_from_note(note),
        "scale": scale,
        "keywords": keywords,
        "traits": derive_traits(name, category, keywords, academic, parent),
        "academic": academic,
        "parent": parent,
        "student": parent,
        "difficulty": None,
        "difficultyStatus": "待补",
        "modules": modules,
        "suitable": unique(suitable)[:5],
        "unsuitable": unique(unsuitable)[:5],
        "careers": careers,
        "graduate": unique(graduate)[:4],
        "universities": universities,
        "risks": unique(risks)[:4],
        "talkTrack": unique(talk_track)[:4],
        "internalNote": "",
        "similar": [],
        "accent": ACCENTS.get(str(identity["disciplineCode"]), "green"),
        "sourceNames": unique(sources),
        "fieldCoverage": {
            "majorDetail": has_major_detail,
            "classDetail": has_class_detail,
            "courses": bool(modules),
            "careers": bool(careers),
            "universities": bool(universities),
        },
    }
    audit_row = {
        "code": code,
        "name": name,
        "discipline": discipline,
        "categoryCode": identity["categoryCode"],
        "category": category,
        "contentStatus": content_status,
        "summaryExact": bool(local),
        "majorTextExact": bool(major_detail),
        "classTextExact": bool(class_detail),
        "courseSheet": course_sheet if modules else "",
        "careerSheet": career_sheet if careers else "",
        "universitySource": university_source if universities else "",
        "sources": " | ".join(unique(sources)),
    }
    return normalized, audit_row


def build_catalog(args) -> tuple[list[dict[str, object]], dict[str, object], dict[str, object], list[dict[str, object]]]:
    official_raw = load_official_records(args.parsed, args.catalog_js)
    official = [official_identity(record) for record in official_raw]
    official_codes = {str(record["code"]) for record in official}
    official_names = {str(record["name"]) for record in official}
    official_classes = {str(record["categoryCode"]) for record in official if record["categoryCode"]}

    local_by_name, summary_audit = load_local_summaries(args.workbook)
    class_details, major_details_raw, class_audit = parse_class_details(args.class_detail)
    course_templates = load_course_templates(args.courses)
    career_templates = load_career_templates(args.careers)
    university_templates = load_university_templates(args.universities)

    # Exact code is primary. A source code typo may be resolved only when the
    # professional name itself exactly matches one and only one official row.
    official_code_by_name = {str(record["name"]): str(record["code"]) for record in official}
    major_details: dict[str, dict[str, object]] = {}
    resolved_name_aliases: list[dict[str, str]] = []
    extra_major_codes: list[str] = []
    for source_code, detail in major_details_raw.items():
        target_code = source_code
        if source_code not in official_codes:
            target_code = official_code_by_name.get(clean_text(detail.get("name")), "")
            if target_code:
                resolved_name_aliases.append(
                    {
                        "sourceCode": source_code,
                        "targetCode": target_code,
                        "name": clean_text(detail.get("name")),
                    }
                )
            else:
                extra_major_codes.append(source_code)
                continue
        major_details[target_code] = detail

    records: list[dict[str, object]] = []
    audit_rows: list[dict[str, object]] = []
    for raw in official_raw:
        identity = official_identity(raw)
        record, audit_row = normalized_record(
            raw,
            local_by_name.get(str(identity["name"])),
            class_details.get(str(identity["categoryCode"])),
            major_details.get(str(identity["code"])),
            course_templates,
            career_templates,
            university_templates,
        )
        records.append(record)
        audit_rows.append(audit_row)

    status_counts = Counter(str(record["contentStatus"]) for record in records)
    field_counts = {
        key: sum(bool(record["fieldCoverage"][key]) for record in records)
        for key in ("majorDetail", "classDetail", "courses", "careers", "universities")
    }
    metadata = {
        "year": 2026,
        "disciplineCount": len({record["disciplineCode"] for record in records}),
        "categoryCount": len({record["categoryCode"] for record in records if record["categoryCode"]}),
        "directCrossDisciplineCount": sum(not record["categoryCode"] for record in records),
        "majorCount": len(records),
        "localSummaryMatchCount": sum(bool(row["summaryExact"]) for row in audit_rows),
        "specificMajorTextMatchCount": sum(bool(row["majorTextExact"]) for row in audit_rows),
        "majorDetailCount": status_counts.get("major", 0),
        "classReferenceCount": status_counts.get("class", 0),
        "catalogOnlyCount": status_counts.get("catalog", 0),
        "fieldCoverage": field_counts,
        "officialNoticeUrl": OFFICIAL_NOTICE_URL,
        "officialPdfUrl": OFFICIAL_PDF_URL,
        "officialSourceName": OFFICIAL_SOURCE_NAME,
        "sourceNames": [
            OFFICIAL_SOURCE_NAME,
            SUMMARY_SOURCE_NAME,
            CLASS_SOURCE_NAME,
            COURSE_SOURCE_NAME,
            CAREER_SOURCE_NAME,
            UNIVERSITY_SOURCE_NAME,
        ],
    }
    report = {
        "metadata": metadata,
        "summaryAudit": {
            **summary_audit,
            "officialExactMatchCount": sum(name in local_by_name for name in official_names),
            "unmatchedSourceNames": sorted(set(local_by_name) - official_names),
        },
        "classDetailAudit": {
            **class_audit,
            "officialClassMatchCount": len(official_classes & set(class_details)),
            "missingOfficialClasses": [
                {
                    "code": code,
                    "name": next(
                        str(record["category"])
                        for record in official
                        if str(record["categoryCode"]) == code
                    ),
                }
                for code in sorted(official_classes - set(class_details))
            ],
            "specificOfficialMajorMatchCount": len(official_codes & set(major_details)),
            "resolvedNameAliases": resolved_name_aliases,
            "extraMajorCodes": sorted(extra_major_codes),
        },
        "templateAudit": {
            "courseSheetsLoaded": sorted(course_templates),
            "careerSheetsLoaded": sorted(career_templates),
            "universitySourcesLoaded": sorted(university_templates),
        },
    }
    return records, metadata, report, audit_rows


def write_javascript(output_path: Path, records: list[dict[str, object]], metadata: dict[str, object]) -> None:
    payload = (
        "(function () {\n"
        "  'use strict';\n\n"
        f"  window.MAJOR_CATALOG_2026_META = {json.dumps(metadata, ensure_ascii=False, indent=2)};\n\n"
        f"  window.MAJOR_CATALOG_2026 = {json.dumps(records, ensure_ascii=False, separators=(',', ':'))};\n"
        "}());\n"
    )
    output_path.write_text(payload, encoding="utf-8")


def write_audit_csv(path: Path, rows: list[dict[str, object]]) -> None:
    fieldnames = list(rows[0]) if rows else []
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser()
    source_group = parser.add_mutually_exclusive_group(required=True)
    source_group.add_argument("--parsed", type=Path)
    source_group.add_argument("--catalog-js", type=Path)
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--class-detail", type=Path)
    parser.add_argument("--courses", type=Path)
    parser.add_argument("--careers", type=Path)
    parser.add_argument("--universities", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--report-json", type=Path)
    parser.add_argument("--report-csv", type=Path)
    args = parser.parse_args()

    records, metadata, report, audit_rows = build_catalog(args)
    if len(records) != 883:
        raise SystemExit(f"Expected 883 majors, got {len(records)}")
    if len({record['code'] for record in records}) != 883:
        raise SystemExit("Duplicate official major codes detected")
    if len({record['id'] for record in records}) != 883:
        raise SystemExit("Duplicate generated major ids detected")
    if metadata["disciplineCount"] != 13:
        raise SystemExit(f"Expected 13 disciplines, got {metadata['disciplineCount']}")
    if metadata["categoryCount"] != 92:
        raise SystemExit(f"Expected 92 major categories, got {metadata['categoryCount']}")
    if metadata["directCrossDisciplineCount"] != 15:
        raise SystemExit(
            "Expected 15 direct majors under interdisciplinary studies, "
            f"got {metadata['directCrossDisciplineCount']}"
        )
    for record in records:
        code = str(record["code"])
        discipline_code = str(record["disciplineCode"])
        category_code = str(record["categoryCode"])
        if not code.startswith(discipline_code):
            raise SystemExit(f"Major {code} does not belong to discipline {discipline_code}")
        if category_code and not code.startswith(category_code):
            raise SystemExit(f"Major {code} does not belong to category {category_code}")
        if not category_code and discipline_code != "14":
            raise SystemExit(f"Only interdisciplinary majors may be directly listed: {code}")

    args.output.parent.mkdir(parents=True, exist_ok=True)
    write_javascript(args.output, records, metadata)
    if args.report_json:
        args.report_json.parent.mkdir(parents=True, exist_ok=True)
        args.report_json.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    if args.report_csv:
        args.report_csv.parent.mkdir(parents=True, exist_ok=True)
        write_audit_csv(args.report_csv, audit_rows)
    print(json.dumps(metadata, ensure_ascii=False))


if __name__ == "__main__":
    main()
