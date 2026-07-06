#!/usr/bin/env python3
"""Build strict plan-category maps from the 312 and expert workbooks.

The frontend uses these maps to classify special/admission channels by exact
subject + batch + school-group code + major code. Group-level maps are emitted
only as a convenience aggregate; major-level maps are the authoritative input.
"""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def code_text(value: Any, width: int | None = None) -> str:
    text = cell_text(value)
    if not text:
        return ""
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    if width and text.isdigit():
        text = text.zfill(width)
    return text


def normalize_batch(value: Any) -> str:
    text = cell_text(value)
    if not text:
        return ""
    if "提前" in text:
        return "提前本科批"
    if "普通批" in text or "本科批" in text or text == "本科院校":
        return "本科批"
    if "专科" in text or "高职" in text:
        return "专科批"
    return text


def guide_bucket(value: str) -> str:
    text = cell_text(value)
    if not text:
        return ""
    if any(x in text for x in ("定向培养军士", "定向军士", "军士生", "士官")):
        return "early-sergeant"
    if any(x in text for x in ("农村订单定向医学生", "医学定向", "免费医学生", "定向医学生", "6医学定向")):
        return "early-medical"
    if any(x in text for x in ("军队院校", "1军校", "军校")):
        return "early-military"
    if any(x in text for x in ("公安政法", "2公安", "公安院校", "公安类")):
        return "early-police"
    if any(x in text for x in ("航海院校", "3航海", "航海")):
        return "early-maritime"
    if any(x in text for x in ("地方专项", "高校专项", "国家专项", "农村专项", "专项计划", "4地方专项", "5高校专项")):
        return "early-special-plan"
    if "强基" in text:
        return "special-strong-base"
    if any(x in text for x in ("8综评A", "综评A", "综合评价A", "综合评价 A")):
        return "special-comprehensive-a"
    if any(x in text for x in ("综评B", "综合评价B", "综合评价 B")):
        return "special-comprehensive-b"
    if any(x in text for x in ("其他院校", "7其他")):
        return "early-other"
    return ""


def map_key(subject: str, batch: str, group_code: str, major_code: str | None = None) -> str:
    parts = [subject, normalize_batch(batch), code_text(group_code, 6)]
    if major_code is not None:
        parts.append(code_text(major_code, 2))
    return "|".join(parts)


def push_group(groups: dict[str, set[str]], subject: str, batch: str, group_code: str, category: str) -> None:
    key = map_key(subject, batch, group_code)
    if category:
        groups[key].add(category)


def flatten_group(groups: dict[str, set[str]]) -> dict[str, str]:
    out: dict[str, str] = {}
    for key, values in groups.items():
        ordered = sorted(values)
        if len(ordered) == 1:
            out[key] = ordered[0]
        elif ordered:
            out[key] = " / ".join(ordered)
    return out


def read_312(path: Path) -> tuple[dict[str, str], dict[str, str], dict[str, str]]:
    major: dict[str, str] = {}
    groups: dict[str, set[str]] = defaultdict(set)
    legacy_group: dict[str, str] = {}
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name, subject in ((" 理科提前批", "物理"), ("文科提前批", "历史")):
            worksheet = workbook[sheet_name]
            rows = worksheet.iter_rows(values_only=True)
            headers = [cell_text(v) for v in next(rows)]
            idx = {name: i for i, name in enumerate(headers) if name}
            group_i = idx["院校组代号"]
            major_i = idx["专业代号"]
            category_i = idx["提前批类别"]
            for row in rows:
                group_code = code_text(row[group_i], 6)
                major_code = code_text(row[major_i], 2)
                category = cell_text(row[category_i])
                if not group_code or not major_code or not category:
                    continue
                major[map_key(subject, "提前本科批", group_code, major_code)] = category
                push_group(groups, subject, "提前本科批", group_code, category)
                legacy_group[f"{subject}|{group_code}"] = category
    finally:
        workbook.close()
    return major, flatten_group(groups), legacy_group


def read_expert(path: Path) -> tuple[dict[str, str], dict[str, str], dict[str, dict[str, str]]]:
    major: dict[str, str] = {}
    groups: dict[str, set[str]] = defaultdict(set)
    labels: dict[str, dict[str, str]] = {}
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook["汇总版"]
        for row in worksheet.iter_rows(values_only=True):
            if len(row) < 13:
                continue
            batch = normalize_batch(row[3])
            subject = cell_text(row[4])
            category = cell_text(row[5])
            school = cell_text(row[7])
            group_code = code_text(row[8], 6)
            group_name = cell_text(row[10])
            major_code = code_text(row[11], 2)
            if subject not in ("物理", "历史") or not batch or not category or not group_code or not major_code:
                continue
            key = map_key(subject, batch, group_code, major_code)
            major[key] = category
            push_group(groups, subject, batch, group_code, category)
            labels[key] = {"school": school, "group": group_name}
    finally:
        workbook.close()
    return major, flatten_group(groups), labels


def crosscheck(
    primary_major: dict[str, str],
    expert_major: dict[str, str],
    expert_labels: dict[str, dict[str, str]],
) -> list[dict[str, str]]:
    conflicts: list[dict[str, str]] = []
    for key, primary_category in sorted(primary_major.items()):
        expert_category = expert_major.get(key)
        if not expert_category:
            continue
        primary_bucket = guide_bucket(primary_category)
        expert_bucket = guide_bucket(expert_category)
        if primary_bucket != expert_bucket:
            label = expert_labels.get(key, {})
            conflicts.append(
                {
                    "key": key,
                    "school": label.get("school", ""),
                    "group": label.get("group", ""),
                    "category312": primary_category,
                    "categoryExpert": expert_category,
                    "bucket312": primary_bucket,
                    "bucketExpert": expert_bucket,
                }
            )
    return conflicts


def write_js(
    output: Path,
    primary_major: dict[str, str],
    primary_group: dict[str, str],
    legacy_group: dict[str, str],
    expert_major: dict[str, str],
    expert_group: dict[str, str],
    conflicts: list[dict[str, str]],
    source_312: Path,
    source_expert: Path,
) -> None:
    payloads = {
        "PLAN_CATEGORY_BY_MAJOR_312": primary_major,
        "PLAN_CATEGORY_BY_GROUP_312": primary_group,
        "EARLY_BATCH_CATEGORIES": legacy_group,
        "PLAN_CATEGORY_BY_MAJOR_EXPERT": expert_major,
        "PLAN_CATEGORY_BY_GROUP_EXPERT": expert_group,
        "PLAN_CATEGORY_CROSSCHECK_CONFLICTS": conflicts,
        "PLAN_CATEGORY_SOURCE_META": {
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "primarySource": source_312.name,
            "secondarySource": source_expert.name,
            "primaryMajorCount": len(primary_major),
            "primaryGroupCount": len(primary_group),
            "expertMajorCount": len(expert_major),
            "expertGroupCount": len(expert_group),
            "conflictCount": len(conflicts),
            "matchKey": "subject|normalizedBatch|schoolGroupCode|majorCode",
            "precedence": "312 major map, then 312 group aggregate, then expert major map, then expert group aggregate",
        },
    }
    lines = [
        "// Generated by scripts/build-plan-category-maps.py.",
        "// Matching key: subject|normalized batch|school group code|major code.",
    ]
    for name, value in payloads.items():
        lines.append(f"window.{name}={json.dumps(value, ensure_ascii=False, separators=(',', ':'))};")
    output.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-312", required=True, type=Path)
    parser.add_argument("--source-expert", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    primary_major, primary_group, legacy_group = read_312(args.source_312)
    expert_major, expert_group, expert_labels = read_expert(args.source_expert)
    conflicts = crosscheck(primary_major, expert_major, expert_labels)
    write_js(
        args.output,
        primary_major,
        primary_group,
        legacy_group,
        expert_major,
        expert_group,
        conflicts,
        args.source_312,
        args.source_expert,
    )
    print(
        "generated",
        args.output,
        {
            "primaryMajor": len(primary_major),
            "primaryGroup": len(primary_group),
            "expertMajor": len(expert_major),
            "expertGroup": len(expert_group),
            "conflicts": len(conflicts),
        },
    )


if __name__ == "__main__":
    main()
